import os
import re
import json
import time
import asyncio
import random
from pathlib import Path
from typing import Optional, Dict, Any, Tuple

import aiohttp
import discord
from discord import app_commands
from discord.ext import commands
from dotenv import load_dotenv

# google sheets
import gspread
from google.oauth2.service_account import Credentials

# local excel fallback (for testing)
import openpyxl

load_dotenv()

# =========================
# ENV
# =========================
DISCORD_TOKEN = os.getenv("DISCORD_TOKEN", "").strip()
GUILD_ID = int(os.getenv("GUILD_ID", "0"))

# verification feature
VERIFIED_ROLE_NAME = os.getenv("VERIFIED_ROLE_NAME", "✅・Verified").strip()
TOTAL_KARMA_THRESHOLD = int(os.getenv("TOTAL_KARMA_THRESHOLD", "800"))
COMMENT_KARMA_THRESHOLD = int(os.getenv("COMMENT_KARMA_THRESHOLD", "500"))

# logs channel fallback by name (optional)
LOG_CHANNEL_NAME = os.getenv("LOG_CHANNEL_NAME", "logs").strip()

# task system defaults
DEFAULT_SHEET_URL = os.getenv("DEFAULT_SHEET_URL", "").strip()
GOOGLE_CREDS_JSON = os.getenv("GOOGLE_CREDS_JSON", "").strip()

if not DISCORD_TOKEN:
    raise RuntimeError("Missing DISCORD_TOKEN in .env")
if not GUILD_ID:
    raise RuntimeError("Missing GUILD_ID in .env")

# =========================
# DISCORD INTENTS
# =========================
intents = discord.Intents.default()
intents.guilds = True
intents.members = True
intents.reactions = True
intents.messages = True

bot = commands.Bot(command_prefix="!", intents=intents)

# =========================
# STORAGE (config + cooldown state)
# =========================
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)


def _cfg_path(guild_id: int) -> Path:
    return DATA_DIR / f"config_{guild_id}.json"


def _state_path(guild_id: int) -> Path:
    return DATA_DIR / f"state_{guild_id}.json"


def load_config(guild_id: int) -> Dict[str, Any]:
    p = _cfg_path(guild_id)
    if not p.exists():
        return {
            "announce_channel_id": None,
            "logs_channel_id": None,
            "reaction_time_sec": 30,
            "ping_role_id": None,
            "cooldown_role_id": None,
            "cooldown_seconds": 2 * 60 * 60,  # default 2h
            "sheet_url": DEFAULT_SHEET_URL or "",
        }
    return json.loads(p.read_text(encoding="utf-8"))


def save_config(guild_id: int, cfg: Dict[str, Any]) -> None:
    _cfg_path(guild_id).write_text(json.dumps(cfg, indent=2), encoding="utf-8")


def load_state(guild_id: int) -> Dict[str, Any]:
    p = _state_path(guild_id)
    if not p.exists():
        return {"cooldowns": {}}  # user_id (str) -> expiry_ts
    return json.loads(p.read_text(encoding="utf-8"))


def save_state(guild_id: int, state: Dict[str, Any]) -> None:
    _state_path(guild_id).write_text(json.dumps(state, indent=2), encoding="utf-8")


# =========================
# CONSTANTS / HELPERS
# =========================
CLAIM_EMOJI = "✅"
USERNAME_RE = re.compile(r"^[A-Za-z0-9_-]{3,20}$")

# create_task run control
RUN_LOCKS: Dict[int, asyncio.Lock] = {}
RUNNING_JOBS: Dict[int, asyncio.Task] = {}
RUN_EVENTS: Dict[int, asyncio.Event] = {}  # set = running, cleared = paused


def lock_for_guild(guild_id: int) -> asyncio.Lock:
    if guild_id not in RUN_LOCKS:
        RUN_LOCKS[guild_id] = asyncio.Lock()
    return RUN_LOCKS[guild_id]


def run_event_for_guild(guild_id: int) -> asyncio.Event:
    if guild_id not in RUN_EVENTS:
        ev = asyncio.Event()
        ev.set()
        RUN_EVENTS[guild_id] = ev
    return RUN_EVENTS[guild_id]


def is_google_sheet_url(url: str) -> bool:
    u = (url or "").strip().lower()
    return u.startswith("http") and "docs.google.com/spreadsheets" in u


def get_role_by_id(guild: discord.Guild, role_id: Optional[int]) -> Optional[discord.Role]:
    if not role_id:
        return None
    return guild.get_role(int(role_id))


async def send_logs(guild: discord.Guild, message: str) -> None:
    """
    Sends logs to:
      1) configured logs_channel_id (from /config_settings), else
      2) fallback channel name (LOG_CHANNEL_NAME, default "logs")
    """
    cfg = load_config(guild.id)

    ch = None
    logs_id = cfg.get("logs_channel_id")
    if logs_id:
        ch = guild.get_channel(int(logs_id))

    if ch is None and LOG_CHANNEL_NAME:
        ch = discord.utils.get(guild.text_channels, name=LOG_CHANNEL_NAME)

    if ch is None:
        return

    try:
        await ch.send(message, allowed_mentions=discord.AllowedMentions.none())
    except Exception:
        pass


def parse_duration_to_seconds(raw: str) -> int:
    """
    Accepts:
      - "2h", "2hr", "2hrs"
      - "120m", "120min", "120mins"
      - "7200s", "7200sec"
      - "2" (plain number -> minutes)
    """
    s = (raw or "").strip().lower()
    if not s:
        return 2 * 60 * 60

    m = re.match(r"^(\d+)\s*(h|hr|hrs|hour|hours)$", s)
    if m:
        return int(m.group(1)) * 3600

    m = re.match(r"^(\d+)\s*(m|min|mins|minute|minutes)$", s)
    if m:
        return int(m.group(1)) * 60

    m = re.match(r"^(\d+)\s*(s|sec|secs|second|seconds)$", s)
    if m:
        return int(m.group(1))

    if s.isdigit():
        return int(s) * 60

    return 2 * 60 * 60


def now_ts() -> int:
    return int(time.time())


def get_cooldown_expiry(guild_id: int, user_id: int) -> int:
    state = load_state(guild_id)
    cd = state.get("cooldowns", {})
    try:
        return int(cd.get(str(user_id), 0) or 0)
    except Exception:
        return 0


def set_cooldown(guild_id: int, user_id: int, expiry_ts: int) -> None:
    state = load_state(guild_id)
    cd = state.setdefault("cooldowns", {})
    cd[str(user_id)] = int(expiry_ts)
    save_state(guild_id, state)


def clear_cooldown(guild_id: int, user_id: int) -> None:
    state = load_state(guild_id)
    cd = state.get("cooldowns", {})
    if str(user_id) in cd:
        del cd[str(user_id)]
        save_state(guild_id, state)


async def is_member_on_cooldown(guild: discord.Guild, member: discord.Member) -> bool:
    """
    FIXED LOGIC (matches your requirement):
    - user is blocked ONLY if they currently HAVE the cooldown role (🔴・Task Holder)
      AND the stored expiry timestamp is still in the future.
    - if they don't have the role, they are allowed (even if stale cooldown state exists).
    """
    cfg = load_config(guild.id)
    cooldown_role = get_role_by_id(guild, cfg.get("cooldown_role_id"))

    # no cooldown role configured => never block
    if cooldown_role is None:
        clear_cooldown(guild.id, member.id)
        return False

    # if user DOES NOT have the role => allow, and clear any stale saved cooldown
    if cooldown_role not in getattr(member, "roles", []):
        clear_cooldown(guild.id, member.id)
        return False

    # user has the role; check expiry
    exp = get_cooldown_expiry(guild.id, member.id)
    if exp <= now_ts():
        # expired but role still on them => remove role and allow
        try:
            await member.remove_roles(cooldown_role, reason="cooldown expired (auto-clean)")
        except Exception:
            pass
        clear_cooldown(guild.id, member.id)
        return False

    return True


# # =========================
# # REDDIT KARMA  (OLD)
# # =========================
# async def fetch_reddit_karma(username: str) -> tuple[int, int, int]:
#     url = f"https://www.reddit.com/user/{username}/about.json"
#     headers = {"User-Agent": "discord-reddit-karma-verify/6.0"}

#     timeout = aiohttp.ClientTimeout(total=10)
#     async with aiohttp.ClientSession(timeout=timeout) as session:
#         async with session.get(url, headers=headers) as resp:
#             if resp.status == 404:
#                 raise ValueError("reddit user not found")
#             if resp.status == 429:
#                 raise ValueError("rate limited by reddit, try again soon")
#             if resp.status != 200:
#                 raise ValueError(f"reddit error http {resp.status}")
#             data = await resp.json()

#     payload = data.get("data", {}) if isinstance(data, dict) else {}
#     if payload.get("is_suspended") is True:
#         raise ValueError("reddit account is suspended")

#     link_karma = int(payload.get("link_karma", 0))
#     comment_karma = int(payload.get("comment_karma", 0))
#     total = link_karma + comment_karma
#     return link_karma, comment_karma, total


# =========================
# REDDIT KARMA (JSON first, HTML fallback)
# =========================
REDDIT_USER_AGENT = os.getenv(
    "REDDIT_USER_AGENT",
    "discord-karma-verifier/1.0 (contact: admin@example.com)"
).strip()

# Reddit proxy settings (optional)
REDDIT_PROXY_URL = os.getenv("REDDIT_PROXY_URL", "").strip()
REDDIT_PROXY_USERNAME = os.getenv("REDDIT_PROXY_USERNAME", "").strip()
REDDIT_PROXY_PASSWORD = os.getenv("REDDIT_PROXY_PASSWORD", "").strip()

# Build proxy URL with embedded auth if configured (URL-encode for special chars like commas)
_REDDIT_PROXY = None
if REDDIT_PROXY_URL and REDDIT_PROXY_USERNAME and REDDIT_PROXY_PASSWORD:
    from urllib.parse import quote
    # URL-encode username and password to handle special characters (commas, colons, etc)
    encoded_user = quote(REDDIT_PROXY_USERNAME, safe="")
    encoded_pass = quote(REDDIT_PROXY_PASSWORD, safe="")
    _REDDIT_PROXY = REDDIT_PROXY_URL.replace("http://", f"http://{encoded_user}:{encoded_pass}@")

# simple in-memory cache to reduce reddit hits
_REDDIT_CACHE: Dict[str, Tuple[int, int, int, int]] = {}  # username -> (post, comment, total, expiry_ts)
_REDDIT_CACHE_TTL = 300  # 5 min


async def fetch_reddit_karma(username: str) -> tuple[int, int, int]:
    """
    returns (link_karma, comment_karma, total)
    tries:
      1) about.json (may 403 from cloud IPs)
      2) old.reddit.com HTML parse fallback (best-effort)
    """
    u = username.strip()

    # cache
    now = int(time.time())
    cached = _REDDIT_CACHE.get(u.lower())
    if cached and now < cached[3]:
        return cached[0], cached[1], cached[2]

    headers = {
        "User-Agent": REDDIT_USER_AGENT,
        "Accept": "application/json,text/html;q=0.9,*/*;q=0.8",
    }

    async def _try_about_json(url: str) -> Optional[tuple[int, int, int]]:
        timeout = aiohttp.ClientTimeout(total=10)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url, headers=headers, proxy=_REDDIT_PROXY) as resp:
                if resp.status == 404:
                    raise ValueError("reddit user not found")
                if resp.status == 429:
                    raise ValueError("rate limited by reddit, try again soon")

                if resp.status == 200:
                    data = await resp.json()
                    payload = data.get("data", {}) if isinstance(data, dict) else {}
                    if payload.get("is_suspended") is True:
                        raise ValueError("reddit account is suspended")

                    link_karma = int(payload.get("link_karma", 0))
                    comment_karma = int(payload.get("comment_karma", 0))
                    total = link_karma + comment_karma
                    return link_karma, comment_karma, total

                # 401/403 often happens on hosted IPs
                if resp.status in (401, 403):
                    return None

                # other errors
                txt = await resp.text()
                raise ValueError(f"reddit error http {resp.status}: {txt[:120]}")

    # 1) try json endpoints
    for url in (
        f"https://www.reddit.com/user/{u}/about.json",
        f"https://old.reddit.com/user/{u}/about.json",
    ):
        got = await _try_about_json(url)
        if got:
            link_karma, comment_karma, total = got
            _REDDIT_CACHE[u.lower()] = (link_karma, comment_karma, total, now + _REDDIT_CACHE_TTL)
            return got

    # 2) fallback: parse old.reddit.com HTML
    html_url = f"https://old.reddit.com/user/{u}/"
    timeout = aiohttp.ClientTimeout(total=10)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.get(html_url, headers=headers, proxy=_REDDIT_PROXY) as resp:
            if resp.status == 404:
                raise ValueError("reddit user not found")
            if resp.status == 429:
                raise ValueError("rate limited by reddit, try again soon")
            if resp.status in (401, 403):
                raise ValueError("reddit blocked this server ip (403). html fallback also blocked.")
            if resp.status != 200:
                txt = await resp.text()
                raise ValueError(f"reddit error http {resp.status}: {txt[:120]}")
            html = await resp.text()

    # suspended check (best-effort)
    if "account has been suspended" in html.lower():
        raise ValueError("reddit account is suspended")

    # old reddit uses "link karma" + "comment karma" text in the sidebar
    link_m = re.search(r"(\d[\d,]*)\s+(?:link|post)\s+karma", html, flags=re.IGNORECASE)
    com_m = re.search(r"(\d[\d,]*)\s+comment\s+karma", html, flags=re.IGNORECASE)

    if not link_m or not com_m:
        raise ValueError("could not parse karma from profile page (layout changed or blocked).")

    link_karma = int(link_m.group(1).replace(",", ""))
    comment_karma = int(com_m.group(1).replace(",", ""))
    total = link_karma + comment_karma

    _REDDIT_CACHE[u.lower()] = (link_karma, comment_karma, total, now + _REDDIT_CACHE_TTL)
    return link_karma, comment_karma, total






# =========================
# SHEET ACCESS (google OR excel)
# =========================
# Uses:
# - Column A (task number) from A2 onwards
# - Column E (assigned user) written with discord username
SheetTask = Tuple[int, str]  # (row_index, task_no)


def _google_client() -> gspread.Client:
    if not GOOGLE_CREDS_JSON:
        raise RuntimeError("GOOGLE_CREDS_JSON missing in .env (set to file path or JSON content).")
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]

    # Support both file path (local) and JSON content (Railway)
    raw = GOOGLE_CREDS_JSON.strip()
    if raw.startswith("{"):
        # Railway: JSON content stored directly in env var
        try:
            info = json.loads(raw)
            creds = Credentials.from_service_account_info(info, scopes=scopes)
        except json.JSONDecodeError as e:
            raise RuntimeError(f"GOOGLE_CREDS_JSON is invalid JSON: {e}")
    else:
        # Local: file path
        creds = Credentials.from_service_account_file(raw, scopes=scopes)

    return gspread.authorize(creds)


def _google_get_next_task(sheet_url: str) -> Optional[SheetTask]:
    gc = _google_client()
    sh = gc.open_by_url(sheet_url)
    ws = sh.get_worksheet(0)

    values = ws.get_all_values()
    if not values or len(values) < 2:
        return None

    for idx, row in enumerate(values[1:], start=2):
        task_no = (row[0] if len(row) > 0 else "").strip()
        assigned = (row[4] if len(row) > 4 else "").strip()
        if task_no and not assigned:
            return (idx, task_no)

    return None


def _google_assign(sheet_url: str, row_index: int, discord_username: str) -> None:
    gc = _google_client()
    sh = gc.open_by_url(sheet_url)
    ws = sh.get_worksheet(0)
    ws.update_cell(row_index, 5, discord_username)  # E = 5


def _excel_get_next_task(xlsx_path: str) -> Optional[SheetTask]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        task_no = ws.cell(r, 1).value
        assigned = ws.cell(r, 5).value
        if task_no is None:
            continue
        task_no_str = str(task_no).strip()
        assigned_str = "" if assigned is None else str(assigned).strip()
        if task_no_str and not assigned_str:
            return (r, task_no_str)
    return None


def _excel_assign(xlsx_path: str, row_index: int, discord_username: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    ws.cell(row_index, 5).value = discord_username
    wb.save(xlsx_path)


async def get_next_task(sheet_url_or_path: str) -> Optional[SheetTask]:
    src = (sheet_url_or_path or "").strip()
    if not src:
        return None
    if is_google_sheet_url(src):
        return await asyncio.to_thread(_google_get_next_task, src)
    return await asyncio.to_thread(_excel_get_next_task, src)


async def assign_task(sheet_url_or_path: str, row_index: int, discord_username: str) -> None:
    src = (sheet_url_or_path or "").strip()
    if not src:
        raise RuntimeError("sheet url/path not set")
    if is_google_sheet_url(src):
        await asyncio.to_thread(_google_assign, src, row_index, discord_username)
        return
    await asyncio.to_thread(_excel_assign, src, row_index, discord_username)


# =========================
# BATCH TASK FETCHING (for multi-claim feature)
# =========================
def _google_get_tasks_batch(sheet_url: str, n: int) -> list[SheetTask]:
    """Fetch up to n unassigned tasks from Google Sheet."""
    gc = _google_client()
    sh = gc.open_by_url(sheet_url)
    ws = sh.get_worksheet(0)
    values = ws.get_all_values()
    result = []

    if not values or len(values) < 2:
        return result

    for idx, row in enumerate(values[1:], start=2):
        task_no = (row[0] if len(row) > 0 else "").strip()
        assigned = (row[4] if len(row) > 4 else "").strip()
        if task_no and not assigned:
            result.append((idx, task_no))
            if len(result) >= n:
                break

    return result


def _excel_get_tasks_batch(xlsx_path: str, n: int) -> list[SheetTask]:
    """Fetch up to n unassigned tasks from Excel."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    result = []

    for r in range(2, ws.max_row + 1):
        task_no = ws.cell(r, 1).value
        assigned = ws.cell(r, 5).value
        if task_no is None:
            continue
        task_no_str = str(task_no).strip()
        assigned_str = "" if assigned is None else str(assigned).strip()
        if task_no_str and not assigned_str:
            result.append((r, task_no_str))
            if len(result) >= n:
                break

    return result


async def get_tasks_batch(sheet_url_or_path: str, n: int) -> list[SheetTask]:
    """Fetch up to n unassigned tasks (Google or Excel)."""
    src = (sheet_url_or_path or "").strip()
    if not src:
        return []
    if is_google_sheet_url(src):
        return await asyncio.to_thread(_google_get_tasks_batch, src, n)
    return await asyncio.to_thread(_excel_get_tasks_batch, src, n)


# =========================
# COOLDOWN SWEEPER
# =========================
async def cooldown_sweeper() -> None:
    await bot.wait_until_ready()
    while True:
        guild = bot.get_guild(GUILD_ID)
        if guild is None:
            await asyncio.sleep(30)
            continue

        cfg = load_config(guild.id)
        cooldown_role = get_role_by_id(guild, cfg.get("cooldown_role_id"))
        if cooldown_role is None:
            await asyncio.sleep(30)
            continue

        state = load_state(guild.id)
        cooldowns = dict(state.get("cooldowns", {}))
        changed = False

        now = now_ts()
        for uid_str, exp in list(cooldowns.items()):
            try:
                uid = int(uid_str)
                exp_int = int(exp or 0)
            except Exception:
                cooldowns.pop(uid_str, None)
                changed = True
                continue

            if exp_int <= now:
                try:
                    member = await guild.fetch_member(uid)  # fetch fresh, not cached
                    if member and cooldown_role in member.roles:
                        await member.remove_roles(cooldown_role, reason="cooldown expired")
                except Exception:
                    pass

                cooldowns.pop(uid_str, None)
                changed = True

        if changed:
            state["cooldowns"] = cooldowns
            save_state(guild.id, state)

        await asyncio.sleep(30)


# =========================
# TASK CLAIM WAIT (pause-aware + cooldown-fixed)
# =========================
async def wait_for_first_valid_claim(
    guild: discord.Guild,
    message: discord.Message,
    reaction_time_sec: int,
    run_event: asyncio.Event,
) -> Optional[discord.Member]:
    remaining = float(reaction_time_sec)

    async def try_remove_reaction(reaction: discord.Reaction, user: discord.User) -> None:
        try:
            await reaction.remove(user)
        except Exception:
            pass

    def check(reaction: discord.Reaction, user: discord.User) -> bool:
        if user.bot:
            return False
        if reaction.message.id != message.id:
            return False
        return str(reaction.emoji) == CLAIM_EMOJI

    while remaining > 0:
        # pause handling: do not tick down timer
        if not run_event.is_set():
            await asyncio.sleep(0.5)
            continue

        tick = min(1.0, remaining)
        start = time.monotonic()

        try:
            reaction, user = await bot.wait_for("reaction_add", timeout=tick, check=check)
        except asyncio.TimeoutError:
            remaining -= (time.monotonic() - start)
            continue

        remaining -= (time.monotonic() - start)

        try:
            member = guild.get_member(user.id) or await guild.fetch_member(user.id)
        except Exception:
            continue

        # cooldown check (fixed)
        if await is_member_on_cooldown(guild, member):
            await try_remove_reaction(reaction, user)
            continue

        return member

    return None


# =========================
# MULTI-CLAIM WINDOW (assign multiple tasks per ping)
# =========================
async def run_multi_assign_window(
    guild: discord.Guild,
    message: discord.Message,
    task_pool: list[SheetTask],
    sheet_src: str,
    reaction_time_sec: int,
    cooldown_role: Optional[discord.Role],
    cooldown_seconds: int,
    run_event: asyncio.Event,
    announce_ch: discord.TextChannel,
) -> int:
    """
    Listen for reactions on a ping message for reaction_time_sec.
    Assign one random task per valid reactor immediately (no waiting).
    Removes assigned tasks from task_pool (mutates in place).
    Returns count of tasks assigned.
    """
    remaining = float(reaction_time_sec)
    assigned_count = 0

    def check(reaction: discord.Reaction, user: discord.User) -> bool:
        if user.bot:
            return False
        if reaction.message.id != message.id:
            return False
        return str(reaction.emoji) == CLAIM_EMOJI

    async def try_remove_reaction(reaction, user):
        try:
            await reaction.remove(user)
        except Exception:
            pass

    while remaining > 0 and task_pool:
        # pause handling
        if not run_event.is_set():
            await asyncio.sleep(0.5)
            continue

        tick = min(1.0, remaining)
        start = time.monotonic()

        try:
            reaction, user = await bot.wait_for("reaction_add", timeout=tick, check=check)
        except asyncio.TimeoutError:
            remaining -= (time.monotonic() - start)
            continue

        remaining -= (time.monotonic() - start)

        # fetch member (ALWAYS fetch fresh to get updated role state)
        try:
            member = await guild.fetch_member(user.id)
        except Exception:
            continue

        # reject: on cooldown (PRIMARY control)
        if await is_member_on_cooldown(guild, member):
            await try_remove_reaction(reaction, user)
            continue

        # pick a random task from pool
        if not task_pool:
            break
        chosen = random.choice(task_pool)
        task_pool.remove(chosen)
        row_index, task_no = chosen

        # write to sheet
        try:
            await assign_task(sheet_src, row_index, member.name)
        except Exception as e:
            await send_logs(guild, f"⚠️ sheet write failed for task #{task_no}: {e}")
            task_pool.append(chosen)  # put it back on failure
            continue

        # apply cooldown
        expiry = now_ts() + cooldown_seconds
        cooldown_added = False
        if cooldown_role is not None:
            try:
                await member.add_roles(cooldown_role, reason="task claimed cooldown")
                cooldown_added = True
            except Exception:
                pass
        if cooldown_added:
            set_cooldown(guild.id, member.id, expiry)
        else:
            clear_cooldown(guild.id, member.id)

        # DM
        dm_ok = True
        try:
            await member.send(f"you got task no. {task_no}\n{sheet_src}")
        except Exception:
            dm_ok = False

        # in-channel reply
        try:
            await announce_ch.send(
                f"{member.mention} claimed **task #{task_no}**.",
                allowed_mentions=discord.AllowedMentions.none(),
            )
        except Exception:
            pass

        await send_logs(
            guild,
            "✅ **task assigned**\n"
            f"- task: #{task_no} (row {row_index})\n"
            f"- discord: {member.mention} (`{member.id}`)\n"
            f"- wrote col E: {member.name}\n"
            f"- dm: {'sent' if dm_ok else 'failed'}\n"
            f"- cooldown role applied: {'yes' if cooldown_added else 'no'}\n"
            f"- cooldown: {cooldown_seconds}s"
        )

        assigned_count += 1

    return assigned_count


# =========================
# TASK RUNNER (batch)
# =========================
async def run_task_batch(guild: discord.Guild, number_of_tasks: int, run_event: asyncio.Event) -> None:
    cfg = load_config(guild.id)

    announce_id = cfg.get("announce_channel_id")
    if not announce_id:
        await send_logs(guild, "⚠️ create_task failed: announce channel not configured.")
        return

    announce_ch = guild.get_channel(int(announce_id))
    if announce_ch is None:
        await send_logs(guild, "⚠️ create_task failed: announce channel missing/unreadable.")
        return

    sheet_src = str(cfg.get("sheet_url") or "").strip()
    if not sheet_src:
        await send_logs(guild, "⚠️ create_task failed: sheet url/path not configured.")
        return

    reaction_time_sec = int(cfg.get("reaction_time_sec", 30))
    ping_role = get_role_by_id(guild, cfg.get("ping_role_id"))
    cooldown_role = get_role_by_id(guild, cfg.get("cooldown_role_id"))
    cooldown_seconds = int(cfg.get("cooldown_seconds", 7200))

    # load all requested tasks upfront
    try:
        task_pool = await get_tasks_batch(sheet_src, number_of_tasks)
    except Exception as e:
        await send_logs(guild, f"⚠️ sheet read failed: {e}")
        return

    if not task_pool:
        await send_logs(guild, "🛑 no unassigned tasks found in sheet.")
        return

    # ping-window loop: keep posting pings until pool is empty
    while task_pool:
        while not run_event.is_set():
            await asyncio.sleep(0.5)

        count = len(task_pool)
        ping_txt = ping_role.mention if ping_role else ""

        # format cooldown in human-readable form
        if cooldown_seconds >= 3600:
            cooldown_str = f"{cooldown_seconds // 3600} hour{'s' if cooldown_seconds // 3600 != 1 else ''}"
        elif cooldown_seconds >= 60:
            cooldown_str = f"{cooldown_seconds // 60} minute{'s' if cooldown_seconds // 60 != 1 else ''}"
        else:
            cooldown_str = f"{cooldown_seconds} second{'s' if cooldown_seconds != 1 else ''}"

        text = (
            f"{ping_txt}\n"
            f"**React to claim a task** {CLAIM_EMOJI}\n\n"
            f"📋 **{count} task{'s' if count != 1 else ''} available**\n"
            f"⏱️ **Next Task Ping in:** {reaction_time_sec} seconds\n"
            f"⏸️ **Cooldown after claim:** {cooldown_str}\n\n"
            f"Tasks are first come, first serve. After reacting, you'll get a DM with the sheet link and your assigned task number. Look for your Discord username in column E.\n\n"
            f"**EVERYONE SHOULD REACT here to claim a task.**"
        ).strip()

        try:
            msg = await announce_ch.send(text, allowed_mentions=discord.AllowedMentions(roles=True))
            await msg.add_reaction(CLAIM_EMOJI)
        except Exception as e:
            await send_logs(guild, f"⚠️ failed to post task ping: {e}")
            await asyncio.sleep(2)
            continue

        await run_multi_assign_window(
            guild, msg, task_pool, sheet_src,
            reaction_time_sec, cooldown_role, cooldown_seconds,
            run_event, announce_ch,
        )

        # clean up the ping message after window closes
        try:
            await msg.delete()
        except Exception:
            pass

        if task_pool:
            await send_logs(guild, f"🔄 {len(task_pool)} task(s) remaining — opening new ping window.")

    await send_logs(guild, "✅ all tasks assigned.")


# =========================
# EVENTS
# =========================
@bot.event
async def on_ready() -> None:
    guild_obj = discord.Object(id=GUILD_ID)
    bot.tree.copy_global_to(guild=guild_obj)
    await bot.tree.sync(guild=guild_obj)

    asyncio.create_task(cooldown_sweeper())

    print(f"Logged in as {bot.user} | commands synced to guild {GUILD_ID}")


# =========================
# COMMAND: reddit verify
# =========================
@bot.tree.command(
    name="reddit_verify_yourself",
    description="Check Reddit karma and grant the verified role when eligible.",
)
@app_commands.describe(username="Reddit username (without /u/)")
async def reddit_verify_yourself(interaction: discord.Interaction, username: str) -> None:
    await interaction.response.defer(ephemeral=True)

    username = username.strip()
    if not USERNAME_RE.match(username):
        await interaction.followup.send("invalid reddit username format.", ephemeral=True)
        return

    if not interaction.guild or not isinstance(interaction.user, discord.Member):
        await interaction.followup.send("run this inside the server (not dms).", ephemeral=True)
        return

    role = discord.utils.get(interaction.guild.roles, name=VERIFIED_ROLE_NAME)
    if role is None:
        await interaction.followup.send(
            f"role '{VERIFIED_ROLE_NAME}' not found. create it or update VERIFIED_ROLE_NAME in .env.",
            ephemeral=True,
        )
        return

    try:
        link_karma, comment_karma, total = await fetch_reddit_karma(username)
    except ValueError as e:
        await interaction.followup.send(f"could not verify: {e}", ephemeral=True)
        return
    except Exception:
        await interaction.followup.send("unexpected error while contacting reddit.", ephemeral=True)
        return

    if not (total > TOTAL_KARMA_THRESHOLD and comment_karma > COMMENT_KARMA_THRESHOLD):
        await interaction.followup.send(
            f"not eligible. u/{username} has {total} total ({link_karma} post / {comment_karma} comment). "
            f"need > {TOTAL_KARMA_THRESHOLD} total and > {COMMENT_KARMA_THRESHOLD} comment karma.",
            ephemeral=True,
        )
        return

    member: discord.Member = interaction.user
    if role in member.roles:
        await interaction.followup.send("already verified.", ephemeral=True)
        return

    try:
        await member.add_roles(role, reason=f"reddit karma gate: u/{username} total={total}, comment={comment_karma}")
    except discord.Forbidden:
        await interaction.followup.send(
            "missing permission to add that role. move the bot role above verified + grant manage roles.",
            ephemeral=True,
        )
        return

    # log verification
    try:
        await send_logs(
            interaction.guild,
            "✅ **verified granted**\n"
            f"- discord: {member.mention} (`{member.id}`)\n"
            f"- reddit: u/{username}\n"
            f"- karma: total={total}, comment={comment_karma}, post={link_karma}\n"
            f"- rule: total > {TOTAL_KARMA_THRESHOLD} and comment > {COMMENT_KARMA_THRESHOLD}",
        )
    except Exception:
        pass

    await interaction.followup.send("✅ verified role granted.", ephemeral=True)


# =========================
# COMMAND: config settings
# =========================
@bot.tree.command(
    name="config_settings",
    description="Configure task delegation (channels, reaction time, ping role, cooldown tag + duration, sheet url).",
)
@app_commands.checks.has_permissions(manage_guild=True)
@app_commands.describe(
    announce_channel="Channel where task posts go",
    logs_channel="Channel where logs go",
    reaction_time_sec="Seconds users have to react (claim window)",
    ping_role="Role to ping for each task post (optional)",
    cooldown_role="Role/tag given to winners, removed after cooldown duration",
    cooldown_duration="Cooldown duration like 2h, 120m, 7200s (plain number = minutes)",
    google_sheet_url="Google Sheet URL (or local xlsx path for testing). If omitted, keeps previous.",
)
async def config_settings(
    interaction: discord.Interaction,
    announce_channel: discord.TextChannel,
    logs_channel: discord.TextChannel,
    reaction_time_sec: app_commands.Range[int, 5, 3600],
    cooldown_role: discord.Role,
    cooldown_duration: str,
    ping_role: Optional[discord.Role] = None,
    google_sheet_url: Optional[str] = None,
) -> None:
    if not interaction.guild:
        await interaction.response.send_message("run this in a server.", ephemeral=True)
        return

    cfg = load_config(interaction.guild.id)

    if google_sheet_url is not None and google_sheet_url.strip():
        cfg["sheet_url"] = google_sheet_url.strip()
    elif not cfg.get("sheet_url"):
        cfg["sheet_url"] = DEFAULT_SHEET_URL or ""

    cfg["announce_channel_id"] = announce_channel.id
    cfg["logs_channel_id"] = logs_channel.id
    cfg["reaction_time_sec"] = int(reaction_time_sec)
    cfg["ping_role_id"] = ping_role.id if ping_role else None

    cfg["cooldown_role_id"] = cooldown_role.id
    cfg["cooldown_seconds"] = int(parse_duration_to_seconds(cooldown_duration))

    save_config(interaction.guild.id, cfg)

    await interaction.response.send_message(
        "✅ config saved.\n"
        f"- announce: {announce_channel.mention}\n"
        f"- logs: {logs_channel.mention}\n"
        f"- reaction time: {reaction_time_sec}s\n"
        f"- ping role: {ping_role.name if ping_role else '(none)'}\n"
        f"- cooldown tag: {cooldown_role.name}\n"
        f"- cooldown: {cfg['cooldown_seconds']}s\n"
        f"- sheet: {cfg.get('sheet_url') or '(not set)'}",
        ephemeral=True,
    )


# =========================
# COMMAND: create_task batch
# =========================
@bot.tree.command(
    name="create_task",
    description="Start delegating tasks. It keeps reposting the current task until someone claims it.",
)
@app_commands.checks.has_permissions(manage_guild=True)
@app_commands.describe(number_of_tasks="How many tasks to assign (sequentially).")
async def create_task(interaction: discord.Interaction, number_of_tasks: app_commands.Range[int, 1, 500]) -> None:
    if not interaction.guild:
        await interaction.response.send_message("run this in a server.", ephemeral=True)
        return

    guild = interaction.guild
    lock = lock_for_guild(guild.id)
    run_event = run_event_for_guild(guild.id)
    run_event.set()  # starting => running

    job = RUNNING_JOBS.get(guild.id)
    if job and not job.done():
        await interaction.response.send_message("a create_task run is already active. use /cancel_task_run first.", ephemeral=True)
        return

    await interaction.response.send_message(f"✅ starting task delegation for {number_of_tasks} tasks.", ephemeral=True)

    async def runner() -> None:
        async with lock:
            await run_task_batch(guild, int(number_of_tasks), run_event)

    RUNNING_JOBS[guild.id] = asyncio.create_task(runner())


# =========================
# COMMAND: pause/resume/cancel task run
# =========================
@bot.tree.command(name="pause_task_run", description="Pause the active task run (timer stops).")
@app_commands.checks.has_permissions(manage_guild=True)
async def pause_task_run(interaction: discord.Interaction) -> None:
    if not interaction.guild:
        await interaction.response.send_message("run this in a server.", ephemeral=True)
        return

    guild_id = interaction.guild.id
    job = RUNNING_JOBS.get(guild_id)
    if not job or job.done():
        await interaction.response.send_message("no active task run found.", ephemeral=True)
        return

    ev = run_event_for_guild(guild_id)
    ev.clear()
    await interaction.response.send_message("⏸️ task run paused.", ephemeral=True)

    try:
        await send_logs(interaction.guild, "⏸️ task run paused.")
    except Exception:
        pass


@bot.tree.command(name="resume_task_run", description="Resume a paused task run.")
@app_commands.checks.has_permissions(manage_guild=True)
async def resume_task_run(interaction: discord.Interaction) -> None:
    if not interaction.guild:
        await interaction.response.send_message("run this in a server.", ephemeral=True)
        return

    guild_id = interaction.guild.id
    job = RUNNING_JOBS.get(guild_id)
    if not job or job.done():
        await interaction.response.send_message("no active task run found.", ephemeral=True)
        return

    ev = run_event_for_guild(guild_id)
    ev.set()
    await interaction.response.send_message("▶️ task run resumed.", ephemeral=True)

    try:
        await send_logs(interaction.guild, "▶️ task run resumed.")
    except Exception:
        pass


# =========================
# INTERNAL helper (NOT a slash command)
# =========================
async def _cancel_task_run(interaction: discord.Interaction) -> None:
    if not interaction.guild:
        # if response already sent, use followup
        if interaction.response.is_done():
            await interaction.followup.send("run this in a server.", ephemeral=True)
        else:
            await interaction.response.send_message("run this in a server.", ephemeral=True)
        return

    guild_id = interaction.guild.id
    job = RUNNING_JOBS.get(guild_id)

    if job and not job.done():
        job.cancel()

    RUNNING_JOBS.pop(guild_id, None)

    # ensure event is set so future runs aren't stuck paused
    run_event_for_guild(guild_id).set()

    # respond safely even if interaction was already responded to
    try:
        if interaction.response.is_done():
            await interaction.followup.send("🛑 task run cancelled.", ephemeral=True)
        else:
            await interaction.response.send_message("🛑 task run cancelled.", ephemeral=True)
    except Exception:
        pass

    try:
        await send_logs(interaction.guild, "🛑 task run cancelled.")
    except Exception:
        pass


@bot.tree.command(name="cancel_task_run", description="Cancel the active task run.")
@app_commands.checks.has_permissions(manage_guild=True)
async def cancel_task_run(interaction: discord.Interaction) -> None:
    await _cancel_task_run(interaction)


# backward-compatible command name
@bot.tree.command(name="stop_create_task", description="Stop/cancel an active create_task run.")
@app_commands.checks.has_permissions(manage_guild=True)
async def stop_create_task(interaction: discord.Interaction) -> None:
    await _cancel_task_run(interaction)



# =========================
# COMMAND: show_config
# =========================
@bot.tree.command(name="show_config", description="Show current task config.")
@app_commands.checks.has_permissions(manage_guild=True)
async def show_config(interaction: discord.Interaction) -> None:
    if not interaction.guild:
        await interaction.response.send_message("run this in a server.", ephemeral=True)
        return

    cfg = load_config(interaction.guild.id)
    announce = interaction.guild.get_channel(cfg.get("announce_channel_id") or 0)
    logs = interaction.guild.get_channel(cfg.get("logs_channel_id") or 0)
    ping_role = get_role_by_id(interaction.guild, cfg.get("ping_role_id"))
    cooldown_role = get_role_by_id(interaction.guild, cfg.get("cooldown_role_id"))

    await interaction.response.send_message(
        "current config:\n"
        f"- announce: {announce.mention if announce else '(not set)'}\n"
        f"- logs: {logs.mention if logs else '(not set)'}\n"
        f"- reaction_time_sec: {cfg.get('reaction_time_sec')}\n"
        f"- ping_role: {ping_role.name if ping_role else '(none)'}\n"
        f"- cooldown_role: {cooldown_role.name if cooldown_role else '(not set)'}\n"
        f"- cooldown_seconds: {cfg.get('cooldown_seconds')}\n"
        f"- sheet_url: {cfg.get('sheet_url') or '(not set)'}",
        ephemeral=True,
    )


# =========================
# RUN
# =========================
bot.run(DISCORD_TOKEN)
